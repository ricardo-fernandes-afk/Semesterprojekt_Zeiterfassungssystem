"""
Modul: Datenexport für TimeArch.

Dieses Modul exportiert Daten aus der Datenbank in eine Excel-Datei. Es unterstützt die Exporte
von Benutzerdaten und Projektdaten und integriert zusätzliche Informationen wie Benutzereinstellungen
oder Projektphasen.

Funktionen:
-----------
- export_to_excel(export_type, identifier): Exportiert Daten basierend auf dem Exporttyp und der ID (Benutzer oder Projekt).
- format_sheet(worksheet, df, start_row=2, apply_filter=False): Wendet Formatierungen auf ein Excel-Arbeitsblatt an.

Verwendung:
-----------
    from feature_export import export_to_excel

    export_to_excel("user", user_id)  # Exportiert Benutzerdaten
    export_to_excel("project", project_number)  # Exportiert Projektdaten
"""

import pandas as pd
from tkinter.filedialog import asksaveasfilename
from tkinter import messagebox
from db.db_connection import create_connection
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

def export_to_excel(export_type, identifier):
    """
    Exportiert Daten aus der Datenbank in eine Excel-Datei.

    Args:
        export_type (str): Typ des Exports ('user' oder 'project').
        identifier (str): Benutzer-ID oder Projektnummer.

    Datenbankabfragen:
    -------------------
    - Ruft Hauptdaten und zusätzliche Informationen ab, abhängig vom Exporttyp.

    Excel-Export:
    --------------
    - Schreibt Daten in eine Excel-Datei mit formatierter Kopfzeile, Summenzeile und Metadatenblatt.
    - Erstellt separate Blätter für Benutzereinstellungen oder Projektphasen.

    Fehlerbehandlung:
    ------------------
    - Zeigt eine Fehlermeldung an, falls die Datenbankabfrage oder der Datei-Export fehlschlägt.
    """
    connection = create_connection()
    if connection:
        cursor = connection.cursor()
        try:
            # SQL-Abfrage basierend auf Export-Typ
            if export_type == "user":
                query = """
                    SELECT 
                        p.project_number AS projektnummer, 
                        p.project_name AS projektname,
                        s.phase_name AS phase,
                        te.hours AS stunden,
                        te.entry_date AS datum,
                        te.activity AS aktivität,
                        te.note AS notiz
                    FROM time_entries te
                    JOIN projects p ON te.project_number = p.project_number
                    LEFT JOIN sia_phases s ON te.phase_id = s.phase_id
                    WHERE te.user_id = %s
                    ORDER BY te.entry_date;
                """

                user_settings_query = """
                    SELECT 
                        username AS benutzername, 
                        role AS rolle, 
                        default_hours_per_day AS sollstunden_pro_Tag,
                        employment_percentage AS stellenprozent,
                        vacation_hours AS ferien,
                        start_date AS startdatum
                    FROM user_settings
                    JOIN users ON users.user_id = user_settings.user_id
                    WHERE users.user_id = %s;
                """

            elif export_type == "project":
                query = """
                    SELECT 
                        u.username AS benutzername,
                        s.phase_name AS phase,
                        te.hours AS stunden,
                        te.entry_date AS datum,
                        te.activity AS aktivität,
                        te.note AS notiz
                    FROM time_entries te
                    JOIN users u ON te.user_id = u.user_id
                    LEFT JOIN sia_phases s ON te.phase_id = s.phase_id
                    WHERE te.project_number = %s
                    ORDER BY te.entry_date;
                """

                project_phases_query = """
                    SELECT 
                        phase_name AS phase,
                        soll_stunden AS sollstunden
                    FROM project_sia_phases
                    WHERE project_number = %s;
                """

                project_users_query = """
                    SELECT 
                        u.username AS benutzername,
                        u.role AS rolle
                    FROM user_projects up
                    JOIN users u ON up.user_id = u.user_id
                    WHERE up.project_number = %s;
                """

            else:
                raise ValueError("Ungültiger Export-Typ.")

            # Hauptdaten abfragen
            cursor.execute(query, (identifier,))
            data = cursor.fetchall()
            columns = [desc[0].lower() for desc in cursor.description]
            df = pd.DataFrame(data, columns=columns)
            
            if "stunden" in df.columns:
                df["stunden"] = pd.to_numeric(df["stunden"], errors="coerce")

            # Zusätzliche Informationen abfragen
            if export_type == "user":
                cursor.execute(user_settings_query, (identifier,))
                user_settings = cursor.fetchone()
                user_settings_columns = [desc[0] for desc in cursor.description]
                user_settings_df = pd.DataFrame([user_settings], columns=user_settings_columns)
                title = f"Benutzer: {user_settings_df.loc[0, 'benutzername']}"

            elif export_type == "project":
                cursor.execute(project_phases_query, (identifier,))
                project_phases = cursor.fetchall()
                project_phases_columns = [desc[0] for desc in cursor.description]
                project_phases_df = pd.DataFrame(project_phases, columns=project_phases_columns)

                cursor.execute(project_users_query, (identifier,))
                project_users = cursor.fetchall()
                project_users_columns = [desc[0] for desc in cursor.description]
                project_users_df = pd.DataFrame(project_users, columns=project_users_columns)

                project_name_query = "SELECT project_name FROM projects WHERE project_number = %s;"
                cursor.execute(project_name_query, (identifier,))
                project_name = cursor.fetchone()[0]
                title = f"Projekt: {identifier} - {project_name}"

            # Datei speichern
            file_path = asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel-Dateien", "*.xlsx")],
                title="Speichern unter..."
            )

            if not file_path:
                return  # Abbrechen

            def format_sheet(worksheet, df, start_row=2, apply_filter=False):
                """
                Wendet Formatierungen auf ein Excel-Arbeitsblatt an.

                Args:
                    worksheet (openpyxl.worksheet.worksheet.Worksheet): Das zu formatierende Arbeitsblatt.
                    df (pandas.DataFrame): Die Daten, die in das Arbeitsblatt geschrieben wurden.
                    start_row (int, optional): Die Zeile, in der die Formatierung beginnt. Standard ist 2.
                    apply_filter (bool, optional): Gibt an, ob ein AutoFilter auf die Kopfzeile angewendet werden soll.

                Formatierungen:
                ----------------
                - Setzt Kopfzeilenfarben und -schriftart.
                - Passt die Spaltenbreiten automatisch an.
                - Fügt bei Bedarf Filter für die Kopfzeilen hinzu.
                """
                # Filter nur auf Hauptblatt anwenden
                if apply_filter:
                    worksheet.auto_filter.ref = f"A{start_row}:{get_column_letter(len(df.columns))}{worksheet.max_row}"

                # Kopfzeilen anpassen
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="0F8100", end_color="0F8100", fill_type="solid")
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet[f"{get_column_letter(col_num)}{start_row}"]
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                for col_num, column_title in enumerate(df.columns, 1):
                    column_width = max(df[column_title].astype(str).map(len).max(), len(column_title)) + 2
                    worksheet.column_dimensions[get_column_letter(col_num)].width = column_width

            # Excel schreiben
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                # Hauptdaten
                df.to_excel(writer, index=False, sheet_name="Daten", startrow=1)
                worksheet = writer.sheets["Daten"]

                # Titel einfügen
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
                title_cell = worksheet.cell(row=1, column=1)
                title_cell.value = title
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal="center")
                
                # Summenzeile mit Excel-Formel hinzufügen
                if "stunden" in df.columns:
                    total_row_index = len(df) + 4
                    total_cell_label = worksheet.cell(row=total_row_index, column=1)
                    total_cell_label.value = "Gesamt"
                    total_cell_label.font = Font(bold=True)

                    hours_column_index = df.columns.get_loc("stunden") + 1
                    total_cell_formula = worksheet.cell(row=total_row_index, column=hours_column_index)
                    total_cell_formula.value = f"=SUM({get_column_letter(hours_column_index)}3:{get_column_letter(hours_column_index)}{total_row_index - 1})"
                    total_cell_formula.font = Font(bold=True)
                    
                format_sheet(worksheet, df, start_row=2, apply_filter=True)

                # Zusätzliche Informationen hinzufügen
                if export_type == "user":
                    user_settings_df.to_excel(writer, index=False, sheet_name="Benutzereinstellungen", startrow=2)
                    format_sheet(writer.sheets["Benutzereinstellungen"], user_settings_df, start_row=2)
                elif export_type == "project":
                    project_phases_df.to_excel(writer, index=False, sheet_name="Projektphasen", startrow=2)
                    format_sheet(writer.sheets["Projektphasen"], project_phases_df, start_row=2)

                    project_users_df.to_excel(writer, index=False, sheet_name="Projektbenutzer", startrow=2)
                    format_sheet(writer.sheets["Projektbenutzer"], project_users_df, start_row=2)

                # Metadaten-Blatt
                metadata = {
                    "Export-Typ": export_type,
                    "Identifikator": identifier,
                    "Anzahl Datensätze": len(df),
                    "Exportdatum": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                metadata_df = pd.DataFrame(metadata.items(), columns=["Attribut", "Wert"])
                metadata_df.to_excel(writer, index=False, sheet_name="Metadaten", startrow=2)
                format_sheet(writer.sheets["Metadaten"], metadata_df, start_row=2)

            messagebox.showinfo("Erfolg", f"Daten erfolgreich exportiert: {file_path}")

        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Exportieren: {e}")

        finally:
            cursor.close()
            connection.close()
